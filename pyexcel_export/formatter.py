import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.copier import WorksheetCopy
from openpyxl.styles import Alignment

import logging
import json
from io import BytesIO
from collections import OrderedDict
from pathlib import Path
import math
import base64
import re
import importlib_resources
from copy import copy

from .serialize import MyEncoder
from .defaults import Meta

debug_logger = logging.getLogger('debug')

DEFAULTS = json.loads(importlib_resources.read_text('pyexcel_export', 'defaults.json'))


class ExcelFormatter:
    def __init__(self, template_file=None):
        """

        :param str|Path|None template_file:
        """
        is_read = False

        if template_file:
            if not isinstance(template_file, Path):
                template_file = Path(template_file)

            if template_file.exists():
                self.styled_wb = openpyxl.load_workbook(str(template_file.absolute()))
                self.to_stylesheets(self.styled_wb)

                is_read = True

        if not is_read:
            self.styled_wb = openpyxl.Workbook()
            self.styled_wb.active.title = '_template'

    @property
    def data(self):
        output = BytesIO()
        self.styled_wb.save(output)

        return output

    @data.setter
    def data(self, _styles):
        if isinstance(_styles, (dict, OrderedDict)):
            _styles = _styles['excel']
        if not isinstance(_styles, BytesIO):
            _styles = BytesIO(base64.b64decode(_styles))

        self.styled_wb = openpyxl.load_workbook(_styles)

    def save(self, raw_data, out_file, meta=None, retain_meta=False):
        """

        :param raw_data:
        :param str|Path out_file:
        :param meta:
        :param retain_meta:
        :return:
        """
        if not isinstance(out_file, Path):
            out_file = Path(out_file)

        if not meta:
            meta = Meta()

        if '_styles' in meta.keys():
            if isinstance(meta['_styles'], (dict, OrderedDict)):
                if 'excel' in meta['_styles'].keys():
                    self.data = meta['_styles']['excel']
                    meta['_styles']['excel'] = self.data
            else:
                self.data = meta['_styles']
                meta['_styles'] = self.data

        if out_file.exists():
            wb = openpyxl.load_workbook(str(out_file.absolute()))
            self.to_stylesheets(wb)
            self.append_styled_sheets(wb)

            original_sheet_names = []
            extraneous_sheet_names = []
        else:
            wb = self.styled_wb
            original_sheet_names = wb.sheetnames
            extraneous_sheet_names = []

        extraneous_sheet_names.append('_template')
        inserted_sheets = []

        if '_meta' in original_sheet_names:
            wb.remove(wb['_meta'])

        self.create_styled_sheet(wb, '_meta', 0)

        meta_matrix = []

        for k, v in meta.excel_matrix:
            if not k.startswith('_'):
                if isinstance(v, (dict, OrderedDict)):
                    v = json.dumps(v, cls=MyEncoder)
                meta_matrix.append([k, v])

        self.fill_matrix(wb['_meta'], meta_matrix, rules=meta, header_row=None)

        if '_meta' in raw_data.keys():
            raw_data.pop('_meta')

        for sheet_name, cell_matrix in raw_data.items():
            if sheet_name not in original_sheet_names:
                self.create_styled_sheet(wb, sheet_name)
                inserted_sheets.append(sheet_name)

            if meta.get('merge_hidden_tables', DEFAULTS['merge_hidden_tables']) in (True, 'true'):
                if not sheet_name.startswith('_'):
                    self.fill_matrix(wb[sheet_name], cell_matrix, rules=meta)
                else:
                    i = 1
                    if '_meta' not in wb.sheetnames:
                        self.create_styled_sheet(wb, '_meta', 0)

                    while True:
                        while True:
                            if not wb['_meta'].cell(row=i, column=1).value:
                                break
                            i += 1

                        if not wb['_meta'].cell(row=i+1, column=1).value:
                            break
                        i += 1

                    matrix = [[sheet_name]]
                    matrix.extend(cell_matrix)
                    self.fill_matrix(wb['_meta'], matrix, start_row=i, rules=meta, header_row=1)
            else:
                self.fill_matrix(wb[sheet_name], cell_matrix, rules=meta)

        for sheet_name in extraneous_sheet_names:
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])

        for sheet_name in wb.sheetnames:
            if self.is_empty_sheet(wb[sheet_name]):
                wb.remove(wb[sheet_name])

        if not retain_meta:
            if '_meta' in wb.sheetnames:
                wb['_meta'].sheet_state = 'hidden'

        wb.save(str(out_file.absolute()))

    @staticmethod
    def create_styled_sheet(wb, sheet_name, pos: int=None):
        wb.create_sheet(sheet_name, pos)
        if '_template' in wb.sheetnames and sheet_name != '_template':
            WorksheetCopy(wb['_template'], wb[sheet_name]).copy_worksheet()
            # wb[sheet_name].copy_worksheet(wb['_template'])

    def append_styled_sheets(self, wb):
        if '_template' not in wb.sheetnames:
            wb.create_sheet('_template')
            if '_template' in self.styled_wb.sheetnames:
                WorksheetCopy(self.styled_wb['_template'], wb['_template']).copy_worksheet()
                # wb['_template'].copy_worksheet(wb['_template'])

        return wb

    @staticmethod
    def to_stylesheets(wb):
        for ws in wb:
            for row in ws:
                for cell in row:
                    cell.value = None

        return wb

    @staticmethod
    def fill_matrix(ws, cell_matrix, start_row=0, rules=None, header_row=0):
        """

        :param ws:
        :param cell_matrix:
        :param start_row:
        :param rules:
        :param int|None header_row:
        :return:
        """
        max_row = ws.max_row

        for row_num, row in enumerate(cell_matrix):
            for col_num, value in enumerate(row):
                if isinstance(value, (dict, OrderedDict)):
                    value = json.dumps(value, cls=MyEncoder)

                ws.cell(column=(col_num + 1),
                        row=(row_num + start_row + 1),
                        value=value)

                if row_num + start_row + 1 > max_row:
                    source_cell = ws.cell(column=(col_num + 1),
                                          row=(row_num + start_row))
                    target_cell = ws.cell(column=(col_num + 1),
                                          row=(row_num + start_row + 1))

                    if source_cell.has_style:
                        target_cell._style = copy(source_cell._style)

        if rules is not None:
            if rules.get('has_header', DEFAULTS['has_header']) in (True, 'true') \
                    and rules.get('freeze_header', DEFAULTS['freeze_header']) in (True, 'true'):
                if ws.title != '_meta':
                    ws.freeze_panes = 'A2'
            if rules.get('col_width_fit_param_keys', DEFAULTS['col_width_fit_param_keys']) in (True, 'true'):
                if ws.title == '_meta':
                    iter_cols = ws.iter_cols()
                    if iter_cols != tuple():
                        width = max([len(str(cell.value)) for cell in next(iter_cols)])
                        ws.column_dimensions['A'].width = width + 2
            if rules.get('col_width_fit_ids', DEFAULTS['col_width_fit_ids']) in (True, 'true'):
                if header_row is not None:
                    for i, header_item in enumerate(cell_matrix[header_row]):
                        if header_item and re.search(r'(?:(?<=^)|(?<=[\s_-]))id(?:(?=$)|(?=[\s_-]))',
                                                     str(header_item), flags=re.IGNORECASE):
                            col_letter = get_column_letter(i + 1)
                            width = max([len(str(cell.value)) for cell in list(ws.iter_cols())[i]])
                            ws.column_dimensions[col_letter].width = width + 2

            iter_rows = ws.iter_rows()
            if iter_rows != tuple():
                col_width = []

                for i in range(len(next(iter_rows))):
                    col_letter = get_column_letter(i + 1)

                    minimum_width = rules.get('minimum_col_width', DEFAULTS['minimum_col_width'])
                    current_width = ws.column_dimensions[col_letter].width
                    if not current_width or current_width < minimum_width:
                        ws.column_dimensions[col_letter].width = minimum_width

                    col_width.append(ws.column_dimensions[col_letter].width)

                for i, row in enumerate(ws):
                    default_height = 12.5

                    multiples_of_font_size = [default_height]
                    for j, cell in enumerate(row):
                        wrap_text = False
                        vertical = None

                        if rules.get('wrap_text', DEFAULTS['wrap_text']):
                            if header_row is not None:
                                wrap_text = True
                            else:
                                continue

                            if cell.value is not None:
                                mul = 0
                                for v in str(cell.value).split('\n'):
                                    mul += math.ceil(len(v) / col_width[j]) * cell.font.size

                                if mul > 0:
                                    multiples_of_font_size.append(mul)

                        if rules.get('align_top', DEFAULTS['align_top']):
                            vertical = "top"

                        cell.alignment = Alignment(wrap_text=wrap_text, vertical=vertical)

                    original_height = ws.row_dimensions[i+1].height
                    if original_height is None:
                        original_height = default_height

                    new_height = max(multiples_of_font_size)
                    max_height = rules.get('maximum_row_height', DEFAULTS['maximum_row_height'])
                    if original_height < new_height or rules.get('reset_height', False):
                        if new_height < max_height:
                            ws.row_dimensions[i + 1].height = new_height
                        else:
                            ws.row_dimensions[i + 1].height = max_height

    @staticmethod
    def is_empty_sheet(ws):
        def is_not_empty():
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        return True

            return False

        return not is_not_empty()
